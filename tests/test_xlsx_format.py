"""Tests for XLSX extractor and sanitizer (tasks 2.2 and 2.3)."""

from __future__ import annotations

import io
import sys
from unittest.mock import MagicMock

import openpyxl
import pytest

from app.models.extraction import ExtractionResult, SpanMap
from app.models.findings import EntityType, Finding, RedactionStyle
from app.pipeline.extractors.spreadsheet import XlsxExtractor
from app.pipeline.sanitizers.spreadsheet import XlsxSanitizer


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """Build an XLSX file in memory from a dict of sheet_name → rows."""
    wb = openpyxl.Workbook()
    # Remove the default sheet created by openpyxl.
    wb.remove(wb.active)

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_with_formula(
    sheets: dict[str, list[list]],
    formulas: dict[str, dict[str, str]],
) -> bytes:
    """Build an XLSX with explicit formula cells.

    *formulas* maps sheet_name → {cell_ref: formula_string}, e.g.
    ``{"Sheet1": {"C2": "=A2&B2"}}``.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
        for cell_ref, formula in formulas.get(name, {}).items():
            ws[cell_ref] = formula

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# XlsxExtractor tests
# ---------------------------------------------------------------------------


class TestXlsxExtractor:
    """Verify XLSX extraction produces header: value text."""

    def test_single_sheet_extraction(self):
        content = _make_xlsx({
            "People": [
                ["Name", "Email"],
                ["Juan Garcia", "juan@test.com"],
            ],
        })
        extractor = XlsxExtractor()
        result = extractor.extract(content, "test.xlsx")

        assert isinstance(result, ExtractionResult)
        assert "--- Sheet: People ---\n" in result.text
        assert "Name: Juan Garcia\n" in result.text
        assert "Email: juan@test.com\n" in result.text
        assert len(result.span_map) == 0
        assert result.pages == 1
        assert result.is_scanned is False

    def test_multi_sheet_extraction(self):
        content = _make_xlsx({
            "Employees": [
                ["Name", "DNI"],
                ["Juan Garcia", "12345678Z"],
            ],
            "Clients": [
                ["Name", "Phone"],
                ["Maria Lopez", "+34 612345678"],
            ],
        })
        extractor = XlsxExtractor()
        result = extractor.extract(content, "multi.xlsx")

        assert "--- Sheet: Employees ---\n" in result.text
        assert "Name: Juan Garcia\n" in result.text
        assert "DNI: 12345678Z\n" in result.text
        assert "--- Sheet: Clients ---\n" in result.text
        assert "Name: Maria Lopez\n" in result.text
        assert "Phone: +34 612345678\n" in result.text
        assert result.pages == 2

    def test_empty_workbook(self):
        content = _make_xlsx({"Empty": []})
        extractor = XlsxExtractor()
        result = extractor.extract(content, "empty.xlsx")

        assert "--- Sheet: Empty ---\n" in result.text
        # No data rows beyond the sheet header.
        assert result.text.strip() == "--- Sheet: Empty ---"
        assert result.pages == 1
        assert result.is_scanned is False

    def test_numeric_cells_are_converted(self):
        content = _make_xlsx({
            "Data": [
                ["ID", "Value"],
                [1, 42.5],
            ],
        })
        extractor = XlsxExtractor()
        result = extractor.extract(content, "nums.xlsx")

        assert "ID: 1\n" in result.text
        assert "Value: 42.5\n" in result.text

    def test_none_cells_are_skipped(self):
        content = _make_xlsx({
            "Sparse": [
                ["Name", "Email"],
                ["Juan Garcia", None],
            ],
        })
        extractor = XlsxExtractor()
        result = extractor.extract(content, "sparse.xlsx")

        assert "Name: Juan Garcia\n" in result.text
        assert "Email:" not in result.text

    def test_images_list_is_empty(self):
        content = _make_xlsx({"A": [["X"], [1]]})
        extractor = XlsxExtractor()
        result = extractor.extract(content, "test.xlsx")

        assert result.images == []

    def test_headers_only_sheet(self):
        """Sheet with only one row (headers, no data) produces no value lines."""
        content = _make_xlsx({"Headers": [["Name", "Email"]]})
        extractor = XlsxExtractor()
        result = extractor.extract(content, "headers_only.xlsx")

        # Only the sheet separator should be present; headers row is consumed
        # as column names but there are no data rows to produce output.
        assert "--- Sheet: Headers ---\n" in result.text
        assert "Name:" not in result.text


# ---------------------------------------------------------------------------
# XlsxSanitizer tests
# ---------------------------------------------------------------------------


class TestXlsxSanitizer:
    """Verify XLSX sanitization replaces PII and preserves structure."""

    def test_basic_redaction_black_style(self):
        content = _make_xlsx({
            "People": [
                ["Name", "Email"],
                ["Juan Garcia", "juan@test.com"],
            ],
        })
        findings = [
            Finding(entity_type=EntityType.PERSON_NAME, original_text="Juan Garcia", score=0.9),
            Finding(entity_type=EntityType.EMAIL, original_text="juan@test.com", score=0.99),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx", style=RedactionStyle.BLACK)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        values = [cell.value for row in ws.iter_rows() for cell in row]

        assert "Juan Garcia" not in values
        assert "juan@test.com" not in values
        assert "\u2588" * 11 in values  # len("Juan Garcia")
        assert "\u2588" * 13 in values  # len("juan@test.com")

    def test_placeholder_style(self):
        content = _make_xlsx({
            "Sheet1": [
                ["Name", "Email"],
                ["Juan Garcia", "juan@test.com"],
            ],
        })
        findings = [
            Finding(entity_type=EntityType.EMAIL, original_text="juan@test.com", score=0.99),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx", style=RedactionStyle.PLACEHOLDER)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        values = [cell.value for row in ws.iter_rows() for cell in row]

        assert "juan@test.com" not in values
        assert "[EMAIL]" in values

    def test_blur_style_same_as_placeholder(self):
        content = _make_xlsx({
            "Sheet1": [
                ["Name", "Phone"],
                ["Juan", "+34 612345678"],
            ],
        })
        findings = [
            Finding(entity_type=EntityType.PHONE, original_text="+34 612345678", score=0.9),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx", style=RedactionStyle.BLUR)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        values = [cell.value for row in ws.iter_rows() for cell in row]

        assert "+34 612345678" not in values
        assert "[PHONE]" in values

    def test_formula_cells_cleared(self):
        content = _make_xlsx_with_formula(
            sheets={"Sheet1": [
                ["First", "Last", "Full"],
                ["Juan", "Garcia", ""],
            ]},
            formulas={"Sheet1": {"C2": '=A2&" "&B2'}},
        )
        findings = [
            Finding(entity_type=EntityType.PERSON_NAME, original_text="Juan", score=0.9),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx", style=RedactionStyle.PLACEHOLDER)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        cell_c2 = ws["C2"]

        # Formula should be cleared and replaced with redaction text.
        assert cell_c2.data_type != "f"
        assert cell_c2.value == "[PERSON_NAME]"

    def test_structure_preserved(self):
        content = _make_xlsx({
            "Data": [
                ["A", "B", "C"],
                ["1", "2", "3"],
                ["4", "5", "6"],
            ],
        })
        findings = [
            Finding(entity_type=EntityType.PERSON_NAME, original_text="2", score=0.9),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx")

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert len(rows) == 3
        assert len(rows[0]) == 3  # 3 columns preserved
        assert rows[0] == ("A", "B", "C")  # Headers unchanged

    def test_non_pii_cells_unchanged(self):
        content = _make_xlsx({
            "Sheet1": [
                ["Name", "City"],
                ["Juan Garcia", "Madrid"],
            ],
        })
        findings = [
            Finding(entity_type=EntityType.PERSON_NAME, original_text="Juan Garcia", score=0.9),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx")

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        values = [cell.value for row in ws.iter_rows() for cell in row]

        assert "Madrid" in values
        assert "Juan Garcia" not in values

    def test_no_findings_returns_valid_xlsx(self):
        content = _make_xlsx({"Sheet1": [["A", "B"], ["1", "2"]]})
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, [], "test.xlsx")

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert len(rows) == 2
        assert rows[0] == ("A", "B")
        assert rows[1] == ("1", "2")

    def test_finding_without_original_text_is_skipped(self):
        content = _make_xlsx({"Sheet1": [["A"], ["1"]]})
        findings = [
            Finding(entity_type=EntityType.FACE, original_text=None, score=0.9),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx")

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws["A2"].value == "1"

    def test_pii_spanning_partial_cell(self):
        """PII that is part of a cell value should only replace the PII portion."""
        content = _make_xlsx({
            "Sheet1": [
                ["Info"],
                ["Contact: juan@test.com today"],
            ],
        })
        findings = [
            Finding(entity_type=EntityType.EMAIL, original_text="juan@test.com", score=0.99),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx", style=RedactionStyle.PLACEHOLDER)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws["A2"].value == "Contact: [EMAIL] today"

    def test_multi_sheet_sanitization(self):
        content = _make_xlsx({
            "Sheet1": [["Name"], ["Juan Garcia"]],
            "Sheet2": [["Name"], ["Juan Garcia"]],
        })
        findings = [
            Finding(entity_type=EntityType.PERSON_NAME, original_text="Juan Garcia", score=0.9),
        ]
        sanitizer = XlsxSanitizer()
        result = sanitizer.sanitize(content, findings, "test.xlsx", style=RedactionStyle.PLACEHOLDER)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        for ws in wb.worksheets:
            assert ws["A2"].value == "[PERSON_NAME]"


# ---------------------------------------------------------------------------
# Import guard tests
# ---------------------------------------------------------------------------


class TestOpenpyxlImportGuard:
    """Verify _require_openpyxl raises RuntimeError when openpyxl is missing."""

    def test_extractor_raises_without_openpyxl(self, monkeypatch):
        import app.pipeline.extractors.spreadsheet as mod

        original_import = __builtins__.__import__ if hasattr(__builtins__, "__import__") else __import__

        def _mock_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return original_import(name, *args, **kwargs)

        monkeypatch.setattr("builtins.__import__", _mock_import)

        with pytest.raises(RuntimeError, match="openpyxl is required"):
            mod._require_openpyxl()

    def test_sanitizer_raises_without_openpyxl(self, monkeypatch):
        """Sanitizer also calls _require_openpyxl (imported from extractors)."""
        import app.pipeline.extractors.spreadsheet as mod

        original_import = __builtins__.__import__ if hasattr(__builtins__, "__import__") else __import__

        def _mock_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return original_import(name, *args, **kwargs)

        monkeypatch.setattr("builtins.__import__", _mock_import)

        with pytest.raises(RuntimeError, match="Install with"):
            mod._require_openpyxl()
